import streamlit as st
from PIL import Image
import openpyxl
import random
import io

def pixelate_image(image, pixel_size=30, threshold=128):
    # 이미지를 L 모드(그레이스케일)로 변환
    img_gray = image.convert("L")
    w, h = img_gray.size
    # 작은 해상도로 줄이고 다시 확대 (픽셀화)
    img_small = img_gray.resize((pixel_size, pixel_size), Image.BICUBIC)
    pixels = img_small.load()

    # 임계값 처리 (이진화)
    for y in range(pixel_size):
        for x in range(pixel_size):
            pixels[x, y] = 0 if pixels[x, y] < threshold else 255

    # 다시 원본 크기로 확대
    img_big = img_small.resize((w, h), Image.NEAREST)
    # 0/255가 들어있는 2차원 배열 만들기
    pixel_map = [[1 if pixels[x, y] == 0 else 0 for x in range(pixel_size)]
                 for y in range(pixel_size)]

    return img_big, pixel_map

def build_puzzle_map_with_horizontal_merge(pixel_map, quizzes):
    if not pixel_map or not quizzes:
        return None

    # O, X 정답 구분
    correct_nums = [num for num, ans in quizzes if ans == "O"]
    incorrect_nums = [num for num, ans in quizzes if ans == "X"]

    # 예외처리 (O, X가 하나도 없는 경우)
    correct_nums = correct_nums or ["??"]
    incorrect_nums = incorrect_nums or ["??"]

    rows = len(pixel_map)
    cols = len(pixel_map[0]) if rows else 0
    puzzle_map = [[None] * cols for _ in range(rows)]

    for r in range(rows):
        for c in range(cols):
            if pixel_map[r][c] == 1:
                puzzle_map[r][c] = random.choice(correct_nums)
            else:
                puzzle_map[r][c] = random.choice(incorrect_nums)
    return puzzle_map

def main():
    st.title("픽셀아트 OX 퀴즈 생성")

    st.subheader("1. 이미지 불러오기")
    uploaded_image = st.file_uploader("이미지 파일 선택", type=["png","jpg","jpeg","bmp","gif"])
    if uploaded_image:
        image = Image.open(uploaded_image)
        st.image(image, caption="원본 이미지", use_column_width=True)

        pixel_size = st.number_input("픽셀 크기", min_value=1, value=30)
        threshold = st.number_input("임계값", min_value=0, max_value=255, value=128)

        if st.button("흑백 픽셀화"):
            img_big, pixel_map = pixelate_image(image, pixel_size, threshold)
            st.image(img_big, caption="흑백 픽셀화 결과", use_column_width=True)
            st.session_state["pixel_map"] = pixel_map
            st.success(f"픽셀화 완료! ({pixel_size}x{pixel_size})")

    st.subheader("2. OX 퀴즈 엑셀 불러오기")
    quiz_file = st.file_uploader("퀴즈 엑셀 선택", type=["xlsx", "xls"])
    if quiz_file:
        wb = openpyxl.load_workbook(quiz_file, data_only=True)
        if "퀴즈" not in wb.sheetnames:
            st.error("'퀴즈' 시트가 없습니다.")
        else:
            ws = wb["퀴즈"]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or len(row) < 2:
                    continue
                num, answer = row[0], row[1]
                if num is None or answer is None:
                    continue
                num = str(num).strip()
                answer = str(answer).strip().upper()
                if answer in ("O", "X"):
                    data.append((num, answer))
            if data:
                st.session_state["quizzes_df"] = data
                st.success(f"퀴즈 {len(data)}개 불러옴")
            else:
                st.error("퀴즈 데이터가 비어있습니다.")

    st.subheader("3. 숨은그림찾기 시트 만들기")
    if "pixel_map" in st.session_state and "quizzes_df" in st.session_state:
        if st.button("시트 만들기"):
            puzzle_map = build_puzzle_map_with_horizontal_merge(
                st.session_state["pixel_map"],
                st.session_state["quizzes_df"]
            )
            if puzzle_map is None:
                st.warning("puzzle_map 생성에 실패했습니다.")
            else:
                # Excel에 쓰기
                out_wb = openpyxl.Workbook()
                ws = out_wb.active
                ws.title = "숨은그림찾기"

                for r, row_data in enumerate(puzzle_map, start=1):
                    for c, val in enumerate(row_data, start=1):
                        # 만약 숫자 변환을 원하면 int() 처리하지만,
                        # 여기서는 그대로 문자열인 경우도 감안
                        try:
                            ws.cell(row=r, column=c, value=int(val))
                        except:
                            ws.cell(row=r, column=c, value=val)

                # 너비 조절
                for col_idx in range(1, len(puzzle_map[0]) + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 3

                # 메모리에 저장한 뒤 다운로드 버튼 제공
                excel_bytes = io.BytesIO()
                out_wb.save(excel_bytes)
                excel_bytes.seek(0)

                st.download_button(
                    label="숨은그림찾기 엑셀 다운로드",
                    data=excel_bytes,
                    file_name="숨은그림찾기.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.write("픽셀화된 이미지와 퀴즈 데이터를 먼저 불러오세요.")

if __name__ == "__main__":
    main()
